#!/usr/bin/python3

# TODO(python): remove python

# Requirements.txt:
#
# # https://pyyaml.org/wiki/PyYAMLDocumentation
# pyaml==23.7.0
# types-PyYAML
#
# # https://pypi.org/project/openpyxl/
# openpyxl==3.1.2
# openpyxl-stubs
# # For getxmp
# defusedxml==0.7.1
#
# # https://jira.readthedocs.io/examples.html
# jira==3.5.2
#


import argparse
import pathlib
import os
import re
import sys
from typing import cast, Any, Optional, List

from jira import JIRA, Issue
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
import yaml

# Thanks to https://stackoverflow.com/a/39317448/1954789

class JJira:
    """ An authentified JIRA client """

    client: JIRA

    def __init__(self, credentials: str):
        # Read the token from jira.yaml
        with open(credentials, "r", encoding='utf-8') as stream:
            config = yaml.safe_load(stream)
            self.client = JIRA(config.get('url'), token_auth=config.get('token'))

    def get_issues(self, jql: str, fields: Optional[List[str]] = None) -> List[Issue]:
        """ List issues and send back a real list

            https://jira.readthedocs.io/examples.html#issues
        """

        issues: List[Issue] = []
        i = 0
        chunk_size = 100
        while True:
            chunk = cast(
                dict[str, Any],
                self.client.search_issues(jql, startAt=i, maxResults=chunk_size, fields = fields)
            )
            i += chunk_size
            issues += cast(List[Issue], chunk.iterable) # type: ignore[attr-defined]
            if i >= chunk.total: # type: ignore[attr-defined]
                break
        return issues

    def print_issues(self, issues: List[Issue]) -> None:
        """ Pretty print issues list
            key: [component,component,...] Summary
        """

        for i in issues:
            print(i.key, end="")
            print(": ", end="")
            if hasattr(i.fields, "components"):
                print(" [", end="")
                print(",".join({ c.name for c in i.fields.components}), end="")
                print("] ", end="")
            print(i.fields.summary, end="")
            print()


def file_latest(in_path: str, globs: List[str]) -> str:
    """ Get the latest file matching globs
    """
    matches: List[pathlib.Path] = []

    for glob in globs:
        print(f"Looking for {glob} in {in_path}")
        matches = matches + list(pathlib.Path(in_path).glob(glob))

    if len(matches) == 0:
        print("Nothing found")
        sys.exit(1)

    return str(max(matches, key=os.path.getctime))

# pylint: disable-next=too-many-arguments
def spreadsheet_open(
                     filename: str,
                     delimiter: str = ',',
                     csv_title: str = 'csv',
                     read_only: bool = False
                     ) -> openpyxl.Workbook:
    """Load anything as a workbook

            can be xls, xlsx, csv
    """
    return openpyxl.load_workbook(filename=filename, read_only=read_only)

def worksheet_headers(worksheet: Worksheet) -> dict[str, int]:
    """ Get the headers of a worksheet

        Returns a map of headers to column number
    """
    row = worksheet['1']
    headers_map = {}
    for cell in row:
        headers_map[cell.value] = cell.column
    return headers_map

def worksheet_fix_header(worksheet: Worksheet, old: str, new: str) -> None:
    """ Change a header to be another value
    """

    currentHeaders = worksheet_headers(worksheet)
    if (currentHeaders.get(new) is None):
        print(f"Converting '{old}' en '{new}' {currentHeaders.get('Heures')}")
        worksheet.cell(row=1, column=currentHeaders.get(old), value=new)

def worksheet_temp() -> Worksheet:
    """ Create a temp spreadsheet in a new workbook
    """
    return openpyxl.Workbook().active

def inject_worksheet_from_array(data: list[dict], target: Worksheet, blanking = 10) -> Worksheet:
    """
        Try to map a sheet onto another

        To make this, it will look to target first row, treating it as headers
        For each of these headers, it look in the original spreadsheet to extract the column

        To allow the spreadsheet to have more columns, it stop at the first empty column (or containing 'x')
    """
    current_row = 2
    for _, line in enumerate(data):
        for col in range (1, target.max_column + 1):
            header = str(target.cell(row = 1, column = col).value)
            if header is None or header == 'x':
                # We stop at the first empty column
                break
            val = line.get(header)
            if val is None:
                val = ''
            target.cell(row = current_row, column = col, value = val)
        current_row += 1

    for _ in range(blanking):
        current_row += 1
        for col in range (1, target.max_column + 1):
            target.cell(row = current_row, column = col, value = '')

    return target

def spreadsheet_split(
        source: Worksheet,
        target: Worksheet,
        key: str,
        value: str,
        split_key: str = "splitted_key",
        split_value: str = "splitted_value",
        regex: str = r"(([A-Z]{3,12}|FM)-?[0-9]{1,4})",
        empty: str = "empty"
        ) -> Worksheet:
    """ Divide each line by looking in dividing_column for regex
        and duplicating it by adding:
         - regex+column
         - target_column
    """
    source_headers = worksheet_headers(source)
    c_key: int = source_headers[key]
    c_value: int = source_headers[value]
    c_split_key: int = len(source_headers) + 1
    c_split_value: int = c_split_key + 1
    compiled_regex: re.Pattern = re.compile(regex)

    nrow = 1
    for row in source.rows:
        # Thanks to https://stackoverflow.com/a/44925460/1954789
        if nrow == 1:
            # Headers
            target.append((cell.value for cell in row))
            target.cell(row=nrow, column=c_split_key, value=split_key)
            target.cell(row=nrow, column=c_split_value, value=split_value)
            nrow += 1
            continue

        if len(row) >= c_key and len(row) >= c_value:
            # Enough data...
            if row[c_key - 1].value:
                matches = compiled_regex.findall(cast(str, row[c_key - 1].value))
                if len(matches) > 0:
                    # Multiple matches
                    hours = float(cast(str, row[c_value - 1].value)) / len(matches)
                    for one_match in matches:
                        target.append((cell.value for cell in row))
                        target.cell(row=nrow, column=c_split_key, value=one_match[0])
                        target.cell(row=nrow, column=c_split_value, value=hours)
                        nrow += 1

                    continue

        # Default case: copy data and add splitted data
        target.append((cell.value for cell in row))
        target.cell(row=nrow, column=c_split_key, value=empty)
        target.cell(row=nrow, column=c_split_value, value=row[c_value - 1].value)
        nrow += 1

    return target

def spreadsheet_inject(
        source: Worksheet,
        target: Worksheet,
        blanking: int = 50
        ) -> Worksheet :
    """
        Try to map a sheet onto another

        To make this, it will look to target first row, treating it as headers
        For each of these headers, it look in the original spreadsheet to extract the column

        To allow the spreadsheet to have more columns, it stop at the first empty column (or containing 'x')
    """

    target_headers_row = target['1']
    source_headers = worksheet_headers(source)

    mappings = {}

    for cell in target_headers_row:
        if cell.value is None or cell.value == 'x':
            # We stop at the first empty column
            break
        if not cell.value in source_headers:
            raise ValueError(f"Target column '{cell.value}' at {cell.column}"
                + f" is not in the source: {source_headers}")
        mappings[cell.column] = source_headers[cell.value]

    current_row = 0
    for row in source.rows:
        current_row += 1
        row_number = row[0].row
        if row_number == 1:
            # Skip header row
            continue

        for m_src, m_target in mappings.items():
            # for cell in row:
            target.cell(row = current_row, column = m_src, value = row[m_target-1].value)

    for _ in range(blanking):
        current_row += 1
        for m_src, _ in mappings.items():
            target.cell(row = current_row, column = m_src, value = '')

    return target

###############################################
#
# Procédure principale (main)
#
###############################################

parser = argparse.ArgumentParser(description='Transform CSV to XLSX')
parser.add_argument('source', default="previous.xlsx", nargs='?')
parser.add_argument('--d365')
parser.add_argument('--target', default='reporting.xlsx')

#
# necessary keys: url, token, email?
# PAT (token): https://jira.blablabla/secure/ViewProfile.jspa?selectedTab=com.atlassian.pats.pats-plugin:jira-user-personal-access-tokens
#
parser.add_argument('--credentials', default=pathlib.Path.join(pathlib.Path.home(), '/restricted/nsi-jira.yaml'))
args = parser.parse_args()

if not args.d365:
    args.d365 = file_latest(str(pathlib.Path(args.source).parent), [ "Transactions de projet_*.xlsx", "Lignes de feuille de temps non validées_*.xlsx" ] )

if not args.target:
    args.target = pathlib.Path(args.source).with_name('reporting.xlsx')

print(f"* Source:    {args.source}")
print(f"* D365:      {args.d365}")
print(f"* Target:    {args.target}")

print("Opening " + args.source)
target = spreadsheet_open(args.source)

print("Opening " + args.d365)
d365 = spreadsheet_open(args.d365) # read_only=True

# Convertion si ce sont des Timesheets en attente au lieu de transactions validées
worksheet_fix_header(d365.active, "Heures", "Quantité")
worksheet_fix_header(d365.active, "Date de N° document", "Date")

print("Splitting by ticket")
splitted = spreadsheet_split(d365.active,
                    worksheet_temp(),
                    key="Commentaire externe",
                    value="Quantité",
                    split_key="Ticket",
                    split_value="Heures",
                    empty="no_ticket",
                    regex=r"((CCFFMG|CCFFMGINCT)-?[0-9]{2,4})"
                    )

try:
    print("Injecting D365")
    spreadsheet_inject(splitted, target['D365'], blanking=100)

    print(f"Connecting to JIRA using {args.credentials}")
    jira = JJira(args.credentials)

    print("Getting Jira tickets")
    issues = jira.get_issues(
        '(project = CCFFMG OR project = "CCFFMG - Gestion des incidents")'
            ' AND (resolutiondate > startofmonth(-1M) OR resolutiondate is EMPTY OR updated > startofmonth(-2M))'
            ' ORDER BY id DESC',
        [ "id", "summary", "components", "labels", "issuetype", "resolution", "created" ])
    # jira.print_issues(issues)

    print("Injecting Jira")
    inject_worksheet_from_array(
        [ {
            "Key": issue.key,
            "Summary": issue.fields.summary,
            "Labels": issue.fields.labels[0] if issue.fields.labels else None,
            "Component": issue.fields.components[0].name if issue.fields.components else None,
            "Type": issue.fields.issuetype.name,
            "Resolution": issue.fields.resolution.name if issue.fields.resolution else None,
            "Creation": issue.fields.created[0:10]
        } for issue in issues ],
        target['Jira'],
        blanking=100)

except KeyboardInterrupt:
    print("Interrupted by user while handling jira")

print("Saving " + str(args.target))
target.save(args.target)
